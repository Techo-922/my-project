from app_init import app
from connect import get_conn
from fastapi import HTTPException,UploadFile, File,Path
from typing import Dict, List,Optional
from pydantic import BaseModel,constr
import openpyxl
from psycopg2 import sql
from fastapi.responses import StreamingResponse
from io import BytesIO

@app.get("/students")
def get_students(page: int = 1, size: int = 20, keyword: Optional[str] = "", department: Optional[str] = "", grade: Optional[str] = "", status: Optional[str] = "",region: Optional[str] = ""):
    conn = get_conn()
    cursor = conn.cursor()

    offset = (page - 1) * size

    query = """
        SELECT s.txx_Sno, s.txx_Sname, s.txx_Sex, s.txx_Age,
               s.txx_DeptNo, d.txx_DeptName, s.txx_ClassID,
               s.txx_StartDate, s.txx_CreditHours,r.txx_regionname
        FROM Tianxx_Students s
        LEFT JOIN Tianxx_Depts d ON s.txx_DeptNo = d.txx_DeptNo
        LEFT JOIN Tianxx_Regions r ON s.txx_RegionID = r.txx_RegionID
        WHERE 1=1
    """

    params = []

    if keyword:
        query += " AND s.txx_Sname ILIKE %s"
        params.append(f"%{keyword}%")

    if department:
        query += " AND s.txx_DeptNo = %s"
        params.append(department)

    if grade:
        query += " AND EXTRACT(YEAR FROM s.txx_StartDate) = %s"
        params.append(int(grade))

    if region:
        query += " AND s.txx_RegionID = %s"
        params.append(region)

    query += " ORDER BY s.txx_Sno LIMIT %s OFFSET %s"
    params.extend([size, offset])

    cursor.execute(query, params)
    rows = cursor.fetchall()

    cursor.close()
    conn.close()

    result = [
        {
            "sno": row[0],
            "sname": row[1],
            "sex": row[2],
            "age": row[3],
            "dept_no": row[4],
            "dept_name": row[5],  # 专业名称
            "class_id": row[6],
            "start_date": row[7],
            "credit_hours": row[8],
            "region_name":row[9]
        }
        for row in rows
    ]

    return {
        "code": 200,
        "message": "查询成功",
        "data": result
    }


class StudentCreate(BaseModel):
    sno: str
    sname: str
    sex: Optional[str] = None
    age: Optional[int] = None
    dept_no: Optional[str] = None
    class_id: Optional[str] = None
    start_date: Optional[str] = None  # 格式 yyyy-mm-dd
    credit_hours: Optional[int] = 0

#添加学生信息
@app.post("/students_add")
def add_student(student: StudentCreate):
    conn = get_conn()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO Tianxx_Students (
                txx_Sno, txx_Sname, txx_Sex, txx_Age, txx_DeptNo,
                txx_ClassID, txx_StartDate, txx_CreditHours
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            student.sno,
            student.sname,
            student.sex,
            student.age,
            student.dept_no,
            student.class_id,
            student.start_date,
            student.credit_hours
        ))
        conn.commit()
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=400, detail=f"添加学生失败: {e}")
    finally:
        cursor.close()
        conn.close()
    return {"code": 200, "message": "添加学生成功"}

@app.post("/students_import")
async def import_students(file: UploadFile = File(...)):
    wb = openpyxl.load_workbook(file.file)
    sheet = wb.active

    conn = get_conn()
    cursor = conn.cursor()

    inserted = 0
    skipped = 0

    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        sno, sname, sex, age, dept_no, class_id, start_date, credit_hours = row

        # 先查学号是否存在
        cursor.execute("SELECT 1 FROM Tianxx_Students WHERE txx_Sno = %s", (sno,))
        if cursor.fetchone():
            # 已存在，跳过
            skipped += 1
            continue

        # 不存在，插入
        cursor.execute("""
            INSERT INTO Tianxx_Students
            (txx_Sno, txx_Sname, txx_Sex, txx_Age, txx_DeptNo, txx_ClassID, txx_StartDate, txx_CreditHours)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (sno, sname, sex, age, dept_no, class_id, start_date, credit_hours))

        inserted += 1

    conn.commit()
    cursor.close()
    conn.close()

    return {
        "code": 200,
        "success": True,
        "message": f"导入完成，成功插入{inserted}条，跳过{skipped}条重复数据",
    }

@app.get("/students_courses/{sno}")
def get_student_courses(sno: str):
    conn = get_conn()
    cursor = conn.cursor()

    query = """
        SELECT 
            c.txx_Cno,           -- 课程编号
            c.txx_Cname,         -- 课程名称
            c.txx_CreditHour,    -- 学分
            t.txx_Tname,         -- 教师姓名
            cl.txx_ClassName,    -- 班级名称
            tc.txx_Term,         -- 学期
            r.txx_Grade          -- 成绩
        FROM Tianxx_Reports r
        JOIN Tianxx_Teaching tc ON r.txx_TeachingID = tc.txx_TeachingID
        JOIN Tianxx_Courses c ON tc.txx_Cno = c.txx_Cno
        JOIN Tianxx_Teachers t ON tc.txx_Tno = t.txx_Tno
        JOIN Tianxx_Classes cl ON tc.txx_ClassID = cl.txx_ClassID
        WHERE r.txx_Sno = %s
        ORDER BY tc.txx_Term DESC
    """

    cursor.execute(query, (sno,))
    rows = cursor.fetchall()
    cursor.close()
    conn.close()

    result = [
        {
            "course_id": row[0],
            "course_name": row[1],
            "credit_hour": row[2],
            "teacher_name": row[3],
            "class_name": row[4],
            "term": row[5],
            "grade": row[6]
        }
        for row in rows
    ]

    return {
        "code": 200,
        "message": "查询成功",
        "data": result
    }

@app.get("/students_grades/{sno}")
def get_student_grades(sno: str):
    conn = get_conn()
    cursor = conn.cursor()

    query = """
        SELECT 
            c.txx_Cno, c.txx_Cname, c.txx_CreditHour,
            t.txx_Tname, teach.txx_Term,
            r.txx_Grade,teach.txx_classid
        FROM Tianxx_Reports r
        JOIN Tianxx_Teaching teach ON r.txx_TeachingID = teach.txx_TeachingID
        JOIN Tianxx_Courses c ON teach.txx_Cno = c.txx_Cno
        JOIN Tianxx_Teachers t ON teach.txx_Tno = t.txx_Tno
        WHERE r.txx_Sno = %s
        ORDER BY teach.txx_Term DESC
    """
    cursor.execute(query, (sno,))
    rows = cursor.fetchall()
    cursor.close()
    conn.close()

    grades = [
        {
            "course_id": row[0],
            "course_name": row[1],
            "credit_hour": row[2],
            "teacher_name": row[3],
            "term": row[4],
            "grade": float(row[5]) if row[5] is not None else None,
            "class_name":row[6]
        }
        for row in rows
    ]

    return {
        "code": 200,
        "data": grades
    }


@app.post("/students_edit")
async def edit_student(student: dict):
    sno = student.get("sno")
    # 其它字段同样获取

    # 执行数据库更新逻辑，例如
    conn = get_conn()
    cursor = conn.cursor()

    try:
        update_query = sql.SQL("""
            UPDATE Tianxx_Students
            SET txx_Sname = %s,
                txx_Sex = %s,
                txx_Age = %s,
                txx_DeptNo = %s,
                txx_ClassID = %s,
                txx_StartDate = %s,
                txx_CreditHours = %s
            WHERE txx_Sno = %s
        """)

        cursor.execute(update_query, (
            student["sname"],
            student["sex"],
            student["age"],
            student["dept_no"],
            student["class_id"],
            student["start_date"],
            student["credit_hours"],
            student["sno"]
        ))

        conn.commit()
        return {"code": 200, "message": "修改成功"}
    except Exception as e:
        conn.rollback()
        return {"code": 500, "message": str(e)}
    finally:
        cursor.close()
        conn.close()


@app.delete("/students_delete/{sno}")
def delete_student(sno: str):
    try:
        conn = get_conn()
        cursor = conn.cursor()

        # 先检查是否存在
        cursor.execute("SELECT 1 FROM Tianxx_Students WHERE txx_Sno = %s", (sno,))
        if cursor.fetchone() is None:
            return {"code": 404, "message": "学生不存在"}

        # 删除记录
        cursor.execute("DELETE FROM Tianxx_Students WHERE txx_Sno = %s", (sno,))
        conn.commit()

        cursor.close()
        conn.close()
        return {"code": 200, "message": "删除成功"}

    except Exception as e:
        print("删除失败：", e)
        return {"code": 500, "message": "服务器内部错误"}


@app.get("/students/export")
def export_students():
    try:
        conn = get_conn()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT
                s.txx_Sno, s.txx_Sname, s.txx_Sex, s.txx_Age,
                r.txx_RegionName, d.txx_DeptName, s.txx_ClassID,
                s.txx_StartDate, s.txx_CreditHours
            FROM Tianxx_Students s
            LEFT JOIN Tianxx_Regions r ON s.txx_RegionID = r.txx_RegionID
            LEFT JOIN Tianxx_Depts d ON s.txx_DeptNo = d.txx_DeptNo
        """)
        rows = cursor.fetchall()

        # 创建 Excel 工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "学生数据"

        headers = ["学号", "姓名", "性别", "年龄", "生源地", "院系", "班级", "入学日期", "已修学分"]
        ws.append(headers)

        for row in rows:
            ws.append(list(row))

        # 保存到内存流
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=students.xlsx"
            }
        )
    except Exception as e:
        print("导出学生数据失败：", e)
        return {"code": 500, "message": "导出失败"}

from fastapi import HTTPException
from connect import get_conn


@app.get("/students_dashboard/{sno}")
def get_student_dashboard(sno: str):
    conn = get_conn()
    cursor = conn.cursor()

    try:
        # 1. 查询学生基本信息和累计学分
        cursor.execute("""
            SELECT s.txx_Sname, COALESCE(SUM(r.txx_Grade * c.txx_CreditHour)/NULLIF(SUM(c.txx_CreditHour),0),0)
            FROM Tianxx_Students s
            LEFT JOIN Tianxx_Reports r ON s.txx_Sno = r.txx_Sno
            LEFT JOIN Tianxx_Teaching t ON r.txx_TeachingID = t.txx_TeachingID
            LEFT JOIN Tianxx_Courses c ON t.txx_Cno = c.txx_Cno
            WHERE s.txx_Sno = %s
            GROUP BY s.txx_Sname
        """, (sno,))
        res = cursor.fetchone()
        if not res:
            raise HTTPException(status_code=404, detail="学生不存在")

        student_name = res[0]
        gpa = float(res[1]) if res[1] is not None else 0.0  # 更安全的 GPA 获取方式

        # 2. 总学分：已修学分（假设成绩>=60的课程学分计入已修）
        cursor.execute("""
            SELECT COALESCE(SUM(c.txx_CreditHour), 0)
            FROM Tianxx_Reports r
            JOIN Tianxx_Teaching t ON r.txx_TeachingID = t.txx_TeachingID
            JOIN Tianxx_Courses c ON t.txx_Cno = c.txx_Cno
            WHERE r.txx_Sno = %s AND r.txx_Grade >= 60
        """, (sno,))
        total_credits = cursor.fetchone()[0]

        # 3. 当前学期数据
        cursor.execute("""
            SELECT
                teach.txx_Term,
                COUNT(DISTINCT c.txx_Cno),
                COALESCE(AVG(r.txx_Grade), 0),
                COALESCE(SUM(c.txx_CreditHour), 0)
            FROM Tianxx_Reports r
            JOIN Tianxx_Teaching teach ON r.txx_TeachingID = teach.txx_TeachingID
            JOIN Tianxx_Courses c ON teach.txx_Cno = c.txx_Cno
            WHERE r.txx_Sno = %s
            GROUP BY teach.txx_Term
            ORDER BY teach.txx_Term DESC
            LIMIT 1
        """, (sno,))
        term_info = cursor.fetchone()
        if term_info:
            current_term, current_term_courses, current_term_avg_grade, current_term_credits = term_info
        else:
            current_term, current_term_courses, current_term_avg_grade, current_term_credits = "", 0, None, 0

        # 4. 成绩趋势（近 6 学期）
        cursor.execute("""
            SELECT teach.txx_Term, AVG(r.txx_Grade)
            FROM Tianxx_Reports r
            JOIN Tianxx_Teaching teach ON r.txx_TeachingID = teach.txx_TeachingID
            WHERE r.txx_Sno = %s
            GROUP BY teach.txx_Term
            ORDER BY teach.txx_Term DESC
            LIMIT 6
        """, (sno,))
        grade_trend_rows = cursor.fetchall()
        grade_trend = [
            {
                "term": row[0],
                "avg_grade": round(float(row[1]), 2) if row[1] is not None else None
            } for row in reversed(grade_trend_rows)
        ]

        # 5. 学分趋势（近 6 学期）
        cursor.execute("""
            SELECT teach.txx_Term, SUM(c.txx_CreditHour)
            FROM Tianxx_Reports r
            JOIN Tianxx_Teaching teach ON r.txx_TeachingID = teach.txx_TeachingID
            JOIN Tianxx_Courses c ON teach.txx_Cno = c.txx_Cno
            WHERE r.txx_Sno = %s
            GROUP BY teach.txx_Term
            ORDER BY teach.txx_Term DESC
            LIMIT 6
        """, (sno,))
        credits_trend_rows = cursor.fetchall()
        credits_trend = [
            {
                "semester": row[0],
                "credits": int(row[1]) if row[1] is not None else 0
            } for row in reversed(credits_trend_rows)
        ]

        # 6. 当前学期课程成绩分布
        cursor.execute("""
            SELECT c.txx_Cname, r.txx_Grade
            FROM Tianxx_Reports r
            JOIN Tianxx_Teaching teach ON r.txx_TeachingID = teach.txx_TeachingID
            JOIN Tianxx_Courses c ON teach.txx_Cno = c.txx_Cno
            WHERE r.txx_Sno = %s AND teach.txx_Term = %s
        """, (sno, current_term))
        score_distribution_rows = cursor.fetchall()
        score_distribution = [
            {
                "course": row[0],
                "score": float(row[1]) if row[1] is not None else None
            } for row in score_distribution_rows
        ]

        return {
            "code": 200,
            "message": "查询成功",
            "data": {
                "sno": sno,
                "name": student_name,
                "total_credits": int(total_credits),
                "current_term": current_term,
                "current_term_courses": current_term_courses,
                "current_term_avg_grade": round(float(current_term_avg_grade), 2) if current_term_avg_grade is not None else None,
                "grade_trend": grade_trend,
                "credits_trend": credits_trend,
                "score_distribution": score_distribution
            }
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()



from pydantic import BaseModel, constr, Field

class StudentProfileUpdate(BaseModel):
    phone: constr(pattern=r'^1[3-9]\d{9}$') = Field(..., description="手机号")
    address: constr(min_length=3, max_length=200) = Field(..., description="地址")


@app.get("/students_info/{sno}")
def get_student_profile(sno: str = Path(..., min_length=1)):
    conn = get_conn()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """
            SELECT txx_Sno, txx_Sname, txx_Sex, txx_Phone, txx_Address
            FROM Tianxx_Students
            WHERE txx_Sno = %s
            """,
            (sno,)
        )
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="学生不存在")
        data = {
            "sno": row[0],
            "name": row[1],
            "gender": row[2],
            "phone": row[3] or "",
            "address": row[4] or ""
        }
        return {"code": 200, "data": data}
    finally:
        cursor.close()
        conn.close()


@app.post("/students/{sno}/update_profile")
def update_student_profile(sno: str, update: StudentProfileUpdate):
    conn = get_conn()
    cursor = conn.cursor()
    try:
        # 检查是否存在该学生
        cursor.execute(
            "SELECT 1 FROM Tianxx_Students WHERE txx_Sno = %s",
            (sno,)
        )
        if cursor.fetchone() is None:
            raise HTTPException(status_code=404, detail="学生不存在")

        # 更新手机号和地址
        cursor.execute(
            """
            UPDATE Tianxx_Students
            SET txx_Phone = %s, txx_Address = %s
            WHERE txx_Sno = %s
            """,
            (update.phone, update.address, sno)
        )
        conn.commit()
        return {"code": 200, "message": "更新成功"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"更新失败: {str(e)}")
    finally:
        cursor.close()
        conn.close()


from fastapi import APIRouter, HTTPException
from pydantic import BaseModel

class CourseSelection(BaseModel):
    teaching_id: int

@app.post("/students/{sno}/select_course")
def select_course(sno: str, selection: CourseSelection):
    conn = get_conn()
    cursor = conn.cursor()

    # 检查是否已经选过
    cursor.execute(
        "SELECT 1 FROM Tianxx_Reports WHERE txx_Sno = %s AND txx_TeachingID = %s",
        (sno, selection.teaching_id)
    )
    if cursor.fetchone():
        cursor.close()
        conn.close()
        raise HTTPException(status_code=400, detail="已选过该课程")

    # 插入选课记录
    cursor.execute(
        "INSERT INTO Tianxx_Reports (txx_Sno, txx_TeachingID, txx_Grade) VALUES (%s, %s, NULL)",
        (sno, selection.teaching_id)
    )
    conn.commit()
    cursor.close()
    conn.close()

    return {"code": 200, "message": "选课成功"}


@app.get("/students/{sno}/selected_courses")
def get_selected_courses(sno: str):
    conn = get_conn()
    cursor = conn.cursor()

    query = """
        SELECT r.txx_TeachingID, c.txx_Cname, c.txx_CreditHour,
               te.txx_Tname, t.txx_Term, r.txx_Grade
        FROM Tianxx_Reports r
        JOIN Tianxx_Teaching t ON r.txx_TeachingID = t.txx_TeachingID
        JOIN Tianxx_Courses c ON t.txx_Cno = c.txx_Cno
        JOIN Tianxx_Teachers te ON t.txx_Tno = te.txx_Tno
        WHERE r.txx_Sno = %s
        ORDER BY t.txx_Term DESC, c.txx_Cname
    """
    cursor.execute(query, (sno,))
    rows = cursor.fetchall()
    cursor.close()
    conn.close()

    return {
        "code": 200,
        "message": "查询成功",
        "data": [
            {
                "teaching_id": row[0],
                "course_name": row[1],
                "credit": row[2],
                "teacher_name": row[3],
                "term": row[4],
                "grade": float(row[5]) if row[5] is not None else None
            }
            for row in rows
        ]
    }



@app.get("/students/{sno}/grades")
def get_student_grades(sno: str):
    conn = get_conn()
    cursor = conn.cursor()

    query = """
        SELECT 
            c.txx_Cno AS course_id,
            c.txx_Cname AS course_name,
            t.txx_Tno AS teacher_id,
            tea.txx_Tname AS teacher_name,
            r.txx_Grade AS grade,
            t.txx_Term AS term
        FROM Tianxx_Reports r
        JOIN Tianxx_Teaching t ON r.txx_TeachingID = t.txx_TeachingID
        JOIN Tianxx_Courses c ON t.txx_Cno = c.txx_Cno
        JOIN Tianxx_Teachers tea ON t.txx_Tno = tea.txx_Tno
        WHERE r.txx_Sno = %s
        ORDER BY t.txx_Term DESC, c.txx_Cno
    """
    cursor.execute(query, (sno,))
    rows = cursor.fetchall()
    cursor.close()
    conn.close()

    data = [
        {
            "course_id": row[0],
            "course_name": row[1],
            "teacher_id": row[2],
            "teacher_name": row[3],
            "grade": float(row[4]) if row[4] is not None else None,
            "term": row[5]
        }
        for row in rows
    ]

    return {
        "code": 200,
        "message": "查询成功",
        "data": data
    }


@app.get("/regions")
def get_all_regions():
    conn = get_conn()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT txx_RegionID, txx_RegionName FROM Tianxx_Regions")
        results = cursor.fetchall()
        regions = [
            {"region_id": row[0], "region_name": row[1]}
            for row in results
        ]
        return {"code":200,"data": regions}
    except Exception as e:
        #return JSONResponse(status_code=500, content={"error": str(e)})
        return {"code": 500, "message": str(e)}
    finally:
        cursor.close()
        conn.close()


