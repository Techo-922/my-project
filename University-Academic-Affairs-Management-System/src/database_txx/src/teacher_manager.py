from fastapi import UploadFile, File,HTTPException
from pydantic import BaseModel,Field
import openpyxl
from app_init import app
from connect import get_conn
from fastapi import Query
from fastapi.responses import JSONResponse
from typing import Optional,List


@app.get("/teachers")
def get_teachers(
    page: int = 1,
    size: int = 20,
    keyword: str = Query("", description="教师姓名模糊搜索关键词")
):
    conn = get_conn()
    cursor = conn.cursor()

    offset = (page - 1) * size

    # 先查询总数
    count_query = "SELECT COUNT(*) FROM Tianxx_Teachers WHERE 1=1"
    count_params = []
    if keyword:
        count_query += " AND txx_Tname ILIKE %s"
        count_params.append(f"%{keyword}%")
    cursor.execute(count_query, count_params)
    total = cursor.fetchone()[0]

    # 查询数据
    query = """
        SELECT txx_Tno, txx_Tname, txx_Sex, txx_Age, txx_Title, txx_Phone
        FROM Tianxx_Teachers
        WHERE 1=1
    """
    params = []
    if keyword:
        query += " AND txx_Tname ILIKE %s"
        params.append(f"%{keyword}%")
    query += " ORDER BY txx_Tno LIMIT %s OFFSET %s"
    params.extend([size, offset])

    cursor.execute(query, params)
    rows = cursor.fetchall()

    cursor.close()
    conn.close()

    result = [
        {
            "tno": row[0],
            "tname": row[1],
            "sex": row[2],
            "age": row[3],
            "title": row[4],
            "phone": row[5]
        }
        for row in rows
    ]

    return {
        "code": 200,
        "message": "查询成功",
        "data": result,
        "total": total  # 这里加上总数
    }

class Teacher(BaseModel):
    tno: str = Field(..., title="教师工号")
    tname: str = Field(..., title="姓名")
    sex: str = Field(..., title="性别")
    age: int = Field(..., ge=18, le=100, title="年龄")
    title: str = Field(..., title="职称")
    phone: str = Field(..., title="手机号")

@app.post("/teachers_add")
def add_teacher(teacher: Teacher):
    conn = get_conn()
    cursor = conn.cursor()
    try:
        # 先检查工号是否存在，避免重复
        cursor.execute("SELECT 1 FROM Tianxx_Teachers WHERE txx_Tno = %s", (teacher.tno,))
        if cursor.fetchone():
            return {"code": 400, "message": "教师工号已存在"}

        # 插入新教师
        insert_query = """
            INSERT INTO Tianxx_Teachers
            (txx_Tno, txx_Tname, txx_Sex, txx_Age, txx_Title, txx_Phone)
            VALUES (%s, %s, %s, %s, %s, %s)
        """
        cursor.execute(insert_query, (teacher.tno, teacher.tname, teacher.sex, teacher.age, teacher.title, teacher.phone))
        conn.commit()
        return {"code": 200, "message": "添加教师成功"}

    except Exception as e:
        conn.rollback()
        print("添加教师失败:", e)
        return {"code": 500, "message": "服务器内部错误"}
    finally:
        cursor.close()
        conn.close()

@app.post("/teachers_import")
async def import_teachers(file: UploadFile = File(...)):
    wb = openpyxl.load_workbook(file.file)
    sheet = wb.active

    conn = get_conn()
    cursor = conn.cursor()

    inserted = 0
    skipped = 0
    failed = 0

    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(cell is None for cell in row):
            continue

        try:
            tno, tname, sex, age, title, phone = row

            # 工号查重
            cursor.execute("SELECT 1 FROM Tianxx_Teachers WHERE txx_Tno = %s", (tno,))
            if cursor.fetchone():
                skipped += 1
                continue

            cursor.execute("""
                INSERT INTO Tianxx_Teachers
                (txx_Tno, txx_Tname, txx_Sex, txx_Age, txx_Title, txx_Phone)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (tno, tname, sex, age, title, phone))

            inserted += 1

        except Exception as e:
            print(f"插入第 {i} 行失败：{e}")
            failed += 1
            conn.rollback()  # 回滚当前事务避免阻断后续插入
            continue

    conn.commit()
    cursor.close()
    conn.close()

    return {
        "code": 200,
        "success": True,
        "message": f"导入完成：成功 {inserted} 条，跳过重复 {skipped} 条，失败 {failed} 条"
    }


@app.delete("/teachers_delete/{tno}")
def delete_teacher(tno: str):
    conn = get_conn()
    cursor = conn.cursor()

    cursor.execute("DELETE FROM Tianxx_Teachers WHERE txx_Tno = %s", (tno,))
    deleted_count = cursor.rowcount

    conn.commit()
    cursor.close()
    conn.close()

    if deleted_count == 0:
        return {"code": 404, "message": "教师不存在", "success": False}

    return {"code": 200, "message": "删除成功", "success": True}

@app.post("/teachers_edit")
async def edit_teacher(teacher: Teacher):
    try:
        conn = get_conn()
        cursor = conn.cursor()

        sql = """
        UPDATE Tianxx_Teachers
        SET 
            txx_Tname = %s,
            txx_Sex = %s,
            txx_Age = %s,
            txx_Title = %s,
            txx_Phone = %s
        WHERE txx_Tno = %s
        """

        cursor.execute(sql, (
            teacher.tname,
            teacher.sex,
            teacher.age,
            teacher.title,
            teacher.phone,
            teacher.tno
        ))

        conn.commit()
        cursor.close()
        conn.close()

        return {"code": 200, "message": "教师信息更新成功"}

    except Exception as e:
        print("数据库更新错误：", e)
        return JSONResponse(content={"code": 500, "message": "教师信息更新失败"}, status_code=500)


@app.get("/teacher_courses")
def get_teacher_courses(tno: str):
    try:
        conn = get_conn()
        cursor = conn.cursor()

        sql = """
            SELECT
                c.txx_Cno AS course_id,
                c.txx_Cname AS course_name,
                c.txx_CreditHour AS credit_hour,
                c.txx_Hour AS total_hours,
                c.txx_AssessType AS assess_type,
                cls.txx_ClassName AS class_name,
                t.txx_Term AS term
            FROM Tianxx_Teaching t
            LEFT JOIN Tianxx_Courses c ON t.txx_Cno = c.txx_Cno
            LEFT JOIN Tianxx_Classes cls ON t.txx_ClassID = cls.txx_ClassID
            WHERE TRIM(t.txx_Tno) = %s
            ORDER BY t.txx_Term DESC, c.txx_Cno
        """

        cursor.execute(sql, (tno.strip(),))  # 去除前后空格
        rows = cursor.fetchall()

        courses = [
            {
                "course_id": row[0],
                "course_name": row[1],
                "credit_hour": row[2],
                "total_hours": row[3],
                "assess_type": row[4],
                "class_name": row[5],
                "term": row[6],
            }
            for row in rows
        ]

        cursor.close()
        conn.close()
        return {"code": 200, "courses": courses}
    except Exception as e:
        print("查询教师授课课程失败：", e)
        return {"code": 500, "message": "查询失败"}

from fastapi.responses import StreamingResponse
from io import BytesIO
import openpyxl

@app.get("/teachers/export")
def export_teachers():
    try:
        conn = get_conn()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT
                txx_Tno,
                txx_Tname,
                txx_Sex,
                txx_Age,
                txx_Title,
                txx_Phone
            FROM Tianxx_Teachers
        """)
        rows = cursor.fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "教师数据"

        headers = ["工号", "姓名", "性别", "年龄", "职称", "电话"]
        ws.append(headers)

        for row in rows:
            ws.append(list(row))

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=teachers.xlsx"}
        )
    except Exception as e:
        print("导出教师数据失败：", e)
        return {"code": 500, "message": "导出失败"}


@app.get("/teacher/{teacher_no}/dashboard")
def get_teacher_dashboard(teacher_no: str):
    conn = get_conn()
    cur = conn.cursor()
    try:
        # 1. 所授课程数
        cur.execute("""
            SELECT COUNT(DISTINCT course_no) AS course_count
            FROM vw_teacher_course_student_grades
            WHERE teacher_no = %s
        """, (teacher_no,))
        course_count = cur.fetchone()[0]

        # 2. 授课学生总数
        cur.execute("""
            SELECT COUNT(DISTINCT student_no) AS total_students
            FROM vw_teacher_course_student_grades
            WHERE teacher_no = %s
        """, (teacher_no,))
        total_students = cur.fetchone()[0]

        # 3. 已提交成绩课程数
        cur.execute("""
            SELECT COUNT(DISTINCT course_no) AS submitted_courses
            FROM vw_teacher_course_student_grades
            WHERE teacher_no = %s AND grade IS NOT NULL
        """, (teacher_no,))
        submitted_courses = cur.fetchone()[0]

        # 4. 平均成绩
        cur.execute("""
            SELECT AVG(grade) AS avg_score
            FROM vw_teacher_course_student_grades
            WHERE teacher_no = %s AND grade IS NOT NULL
        """, (teacher_no,))
        avg_score = cur.fetchone()[0]
        avg_score = round(avg_score, 2) if avg_score else 0.0

        # 5. 各课程平均成绩
        cur.execute("""
            SELECT course_name, AVG(grade) AS avg_score
            FROM vw_teacher_course_student_grades
            WHERE teacher_no = %s AND grade IS NOT NULL
            GROUP BY course_name
        """, (teacher_no,))
        courses = [{"name": row[0], "avg_score": round(row[1], 2)} for row in cur.fetchall()]

        # 6. 成绩等级分布
        cur.execute("""
            SELECT
                COUNT(CASE WHEN grade >= 90 THEN 1 END) AS excellent,
                COUNT(CASE WHEN grade >= 80 AND grade < 90 THEN 1 END) AS good,
                COUNT(CASE WHEN grade >= 60 AND grade < 80 THEN 1 END) AS pass,
                COUNT(CASE WHEN grade < 60 THEN 1 END) AS fail
            FROM vw_teacher_course_student_grades
            WHERE teacher_no = %s AND grade IS NOT NULL
        """, (teacher_no,))
        dist = cur.fetchone()

        cur.execute("""
                    SELECT
                        teacher_name
                    FROM vw_teacher_course_student_grades
                    WHERE teacher_no = %s
                """, (teacher_no,))
        name = cur.fetchone()[0]

        score_distribution = [
            {"name": "优秀（90-100）", "value": dist[0] or 0},
            {"name": "良好（80-89）", "value": dist[1] or 0},
            {"name": "及格（60-79）", "value": dist[2] or 0},
            {"name": "不及格（<60）", "value": dist[3] or 0},
        ]

        return {
            "course_count": course_count,
            "total_students": total_students,
            "submitted_courses": submitted_courses,
            "avg_score": avg_score,
            "courses": courses,
            "score_distribution": score_distribution,
            "name": name,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()

class TeacherProfileUpdate(BaseModel):
    phone: Optional[str]
    address: Optional[str]

@app.post("/teachers/{tno}/update_profile")
def update_teacher_profile(tno: str, update: TeacherProfileUpdate):
    conn = get_conn()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            UPDATE Tianxx_Teachers
            SET txx_Phone = %s,
                txx_Address = %s
            WHERE txx_Tno = %s
        """, (update.phone, update.address, tno))

        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="教师不存在")

        conn.commit()
        return {"message": "教师资料已更新"}

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        cursor.close()
        conn.close()

@app.get("/teachers/{tno}/courses")
def get_teacher_courses(tno: str):
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT DISTINCT teaching_id, course_no, course_name, term
        FROM vw_teacher_course_student_grades
        WHERE teacher_no = %s
        ORDER BY term DESC
    """, (tno,))
    data = cursor.fetchall()
    cursor.close()
    conn.close()
    return [
        {
          "teaching_id": row[0],
          "course_no": row[1],
          "course_name": row[2],
          "term": row[3]
        }
        for row in data
    ]


@app.get("/teachers/{tno}/course_students")
def get_course_students(tno: str, course_no: str, term: str):
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT student_no, student_name, student_sex, student_phone, student_address, grade
        FROM vw_teacher_course_student_grades
        WHERE teacher_no = %s AND course_no = %s AND term = %s
        ORDER BY student_no
    """, (tno, course_no, term))
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return [
        {
            "student_no": r[0],
            "student_name": r[1],
            "student_sex": r[2],
            "student_phone": r[3],
            "student_address": r[4],
            "grade": float(r[5]) if r[5] is not None else None
        }
        for r in rows
    ]

class GradeSubmission(BaseModel):
    student_no: str
    course_no: str
    term: str
    grade: float

class GradeBatchSubmission(BaseModel):
    grades: List[GradeSubmission]

from fastapi import Body, HTTPException

@app.post("/teachers/{tno}/submit_grade")
def submit_grade(tno: str, data: GradeBatchSubmission = Body(...)):
    conn = get_conn()
    cursor = conn.cursor()

    for grade_item in data.grades:
        sno = grade_item.student_no
        course_no = grade_item.course_no
        term = grade_item.term
        grade = grade_item.grade

        # 验证必要字段
        if not all([sno, course_no, term, grade is not None]):
            raise HTTPException(status_code=400, detail="参数缺失")

        # 查询 teaching_id
        cursor.execute("""
            SELECT txx_TeachingID FROM Tianxx_Teaching
            WHERE txx_Tno = %s AND txx_Cno = %s AND txx_Term = %s
        """, (tno, course_no, term))
        teaching_id_row = cursor.fetchone()
        if not teaching_id_row:
            raise HTTPException(status_code=404, detail=f"找不到教学记录: {course_no} {term}")

        teaching_id = teaching_id_row[0]

        # 判断是否已有记录
        cursor.execute("""
            SELECT 1 FROM Tianxx_Reports
            WHERE txx_Sno = %s AND txx_TeachingID = %s
        """, (sno, teaching_id))
        exists = cursor.fetchone()

        if exists:
            cursor.execute("""
                UPDATE Tianxx_Reports
                SET txx_Grade = %s
                WHERE txx_Sno = %s AND txx_TeachingID = %s
            """, (grade, sno, teaching_id))
        else:
            cursor.execute("""
                INSERT INTO Tianxx_Reports (txx_Sno, txx_TeachingID, txx_Grade)
                VALUES (%s, %s, %s)
            """, (sno, teaching_id, grade))

    conn.commit()
    cursor.close()
    conn.close()

    return {"message": "成绩批量提交成功"}


@app.get("/teachers/{tno}/course_students_scores")
def get_course_students_scores(tno: str, course_no: str, term: str):
    conn = get_conn()
    cursor = conn.cursor()
    # 查询学生成绩和排名
    cursor.execute("""
        SELECT
            student_no,
            student_name,
            grade,
            RANK() OVER (ORDER BY grade DESC NULLS LAST) AS rank
        FROM vw_teacher_course_student_grades
        WHERE teacher_no = %s AND course_no = %s AND term = %s
        ORDER BY rank
    """, (tno, course_no, term))
    students = cursor.fetchall()

    # 查询平均成绩
    cursor.execute("""
        SELECT AVG(grade)
        FROM vw_teacher_course_student_grades
        WHERE teacher_no = %s AND course_no = %s AND term = %s
    """, (tno, course_no, term))
    avg_score = cursor.fetchone()[0] or 0

    cursor.close()
    conn.close()

    return {
        "average_score": round(avg_score, 2),
        "students": [
            {
                "student_no": r[0],
                "student_name": r[1],
                "grade": float(r[2]) if r[2] is not None else None,
                "rank": r[3]
            } for r in students
        ]
    }

